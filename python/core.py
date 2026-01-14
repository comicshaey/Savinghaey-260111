# core.py
# 교육공무직 연차유급휴가 계산 핵심 로직 (최종 코어)

from dataclasses import dataclass
from datetime import date, datetime, timedelta
import math
import pandas as pd
from openpyxl import load_workbook
from typing import Optional, Dict


# =========================
# 유틸
# =========================

def to_date(x) -> Optional[date]:
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip().replace(".", "-").replace("/", "-")
    try:
        y, m, d = map(int, s.split("-"))
        return date(y, m, d)
    except Exception:
        return None


def safe_float(x) -> float:
    try:
        return float(str(x).replace("일", "").strip())
    except Exception:
        return 0.0


# =========================
# 데이터 모델
# =========================

@dataclass
class Employee:
    name: str
    first_hire_date: date
    period_type: str        # SCHOOL_YEAR / CALENDAR_YEAR
    schedule_key: str       # 분모 키


@dataclass
class Period:
    start: date
    end: date


# =========================
# 기준기간
# =========================

def get_period(grant_year: int, period_type: str) -> Period:
    if period_type == "SCHOOL_YEAR":
        start = date(grant_year - 1, 3, 1)
        end = date(grant_year, 3, 1) - timedelta(days=1)
    else:
        start = date(grant_year - 1, 1, 1)
        end = date(grant_year - 1, 12, 31)
    return Period(start, end)


# =========================
# 소정근로일수(분모)
# =========================

SCHEDULE_WORK_DAYS: Dict[str, float] = {
    "FULLTIME_260": 260.0,
    "KINDER_236": 236.0,
}

def scheduled_work_days(schedule_key: str) -> float:
    if schedule_key in SCHEDULE_WORK_DAYS:
        return SCHEDULE_WORK_DAYS[schedule_key]
    return float(schedule_key)


# =========================
# 근무상황목록 파서
# =========================

def read_worklog(xlsx_path: str) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() for c in rows[0]]
    df = pd.DataFrame(rows[1:], columns=header)

    colmap = {}
    for c in df.columns:
        cc = c.replace(" ", "")
        if cc in ("근무일자", "일자"):
            colmap[c] = "work_date"
        elif cc in ("근무상황", "근태"):
            colmap[c] = "status"
        elif cc == "시작일자":
            colmap[c] = "start_date"
        elif cc == "종료일자":
            colmap[c] = "end_date"
        elif cc in ("일수", "사용일수"):
            colmap[c] = "days"

    df = df.rename(columns=colmap)

    for c in ["work_date", "start_date", "end_date"]:
        if c in df.columns:
            df[c] = df[c].apply(to_date)
        else:
            df[c] = None

    df["days"] = df.get("days", 0).apply(safe_float)
    df["status"] = df.get("status", "")

    return df


# =========================
# 출근 / 차감 규칙
# =========================

NON_ATTEND = ["결근", "무급", "무단", "휴직"]
ATTEND_DEEMED = ["연가", "연차", "공가", "경조", "출산", "육아", "병가"]

def is_non_attend(status: str) -> bool:
    for a in ATTEND_DEEMED:
        if a in status:
            return False
    for n in NON_ATTEND:
        if n in status:
            return True
    return False


def calc_non_attend_days(df: pd.DataFrame, period: Period) -> float:
    total = 0.0
    for _, r in df.iterrows():
        if not is_non_attend(str(r["status"])):
            continue
        d = r["work_date"]
        if d and period.start <= d <= period.end:
            total += r["days"] if r["days"] else 1.0
    return total


def breakdown_non_attend_by_status(df: pd.DataFrame, period: Period) -> pd.DataFrame:
    rows = []
    for _, r in df.iterrows():
        if not is_non_attend(str(r["status"])):
            continue
        d = r["work_date"]
        if d and period.start <= d <= period.end:
            rows.append({
                "근무상황": r["status"],
                "일수": r["days"] if r["days"] else 1.0
            })
    if not rows:
        return pd.DataFrame(columns=["근무상황", "일수"])
    return (
        pd.DataFrame(rows)
        .groupby("근무상황", as_index=False)["일수"]
        .sum()
        .sort_values("일수", ascending=False)
    )


# =========================
# 연차 계산
# =========================

def years_of_service(first_hire: date, ref: date) -> int:
    y = ref.year - first_hire.year
    if (ref.month, ref.day) < (first_hire.month, first_hire.day):
        y -= 1
    return y


def normal_entitlement(first_hire: date, grant_year: int) -> float:
    y = years_of_service(first_hire, date(grant_year, 1, 1))
    if y < 1:
        return 0.0
    return min(25, 15 + (y - 1) // 2)


def calculate_annual_leave(emp: Employee, worklog: pd.DataFrame, grant_year: int) -> dict:
    period = get_period(grant_year, emp.period_type)
    denom = scheduled_work_days(emp.schedule_key)
    non_att = calc_non_attend_days(worklog, period)

    attend_rate = max(0.0, (denom - non_att) / denom)
    normal = normal_entitlement(emp.first_hire_date, grant_year)
    granted = normal if attend_rate >= 0.8 else round(normal * attend_rate, 2)

    return {
        "기준기간": f"{period.start} ~ {period.end}",
        "소정근로일수": denom,
        "차감일수": non_att,
        "출근율(%)": round(attend_rate * 100, 2),
        "부여연차": granted,
    }
